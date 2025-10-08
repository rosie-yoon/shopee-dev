# -*- coding: utf-8 -*-
from __future__ import annotations

import re
from io import BytesIO
from typing import List, Optional

import gspread
from gspread.utils import rowcol_to_a1
import pandas as pd
from zipfile import ZipFile as _ZipFile

# 프로젝트 공통 유틸
from .utils_common import open_sheet_by_env, safe_worksheet, with_retry, get_env

# ------------------------------------------------------
# 0) XLSX Sanitize: sheetViews / pane 제거 (네임스페이스 포함)
# ------------------------------------------------------
def _sanitize_xlsx_remove_sheetviews(bio: BytesIO) -> BytesIO:
    """
    XLSX(zip) 내부의 xl/worksheets/sheet*.xml 에서
    <sheetViews> 블록, <pane .../> 태그를 제거해 파싱을 안정화한다.
    """
    bio.seek(0)
    raw = bio.read()
    ib = BytesIO(raw)
    ob = BytesIO()
    modified = False

    re_sheetviews_block = re.compile(r"<(?:\w+:)?sheetViews[\s\S]*?</(?:\w+:)?sheetViews>", re.IGNORECASE)
    re_pane_self = re.compile(r"<(?:\w+:)?pane\b[^>]*/>", re.IGNORECASE)
    re_pane_block = re.compile(r"<(?:\w+:)?pane\b[^>]*>[\s\S]*?</(?:\w+:)?pane>", re.IGNORECASE)

    with _ZipFile(ib, "r") as zin, _ZipFile(ob, "w") as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename.startswith("xl/worksheets/sheet") and info.filename.endswith(".xml"):
                text = data.decode("utf-8", errors="ignore")
                new_text = re_sheetviews_block.sub("", text)
                new_text = re_pane_self.sub("", new_text)
                new_text = re_pane_block.sub("", new_text)
                if new_text != text:
                    modified = True
                    data = new_text.encode("utf-8", errors="ignore")
            zout.writestr(info, data)

    if modified:
        ob.seek(0)
        return ob
    return BytesIO(raw)

# ------------------------------------------------------
# 1) 보이는 행(openpyxl) 우선 파서 + 숨김/높이0/아웃라인 숨김 제외
# ------------------------------------------------------
def _is_row_hidden_extended(ws, r_idx: int) -> bool:
    rd = ws.row_dimensions.get(r_idx)
    if not rd:
        return False
    if getattr(rd, "hidden", False):
        return True
    if getattr(rd, "height", None) == 0:
        return True
    if getattr(rd, "outlineLevel", 0) > 0 and getattr(rd, "hidden", False):
        return True
    return False

def _read_with_openpyxl_visible_only(file_bytes: bytes) -> List[List[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True, keep_links=False)
    # 데이터가 가장 많은 시트를 선택
    target = max(wb.worksheets, key=lambda s: (s.max_row or 0) * (s.max_column or 0))

    max_r = target.max_row or 0
    max_c = target.max_column or 0
    if max_r == 0 or max_c == 0:
        return []

    rows: List[List[str]] = []
    for r_idx, row in enumerate(
        target.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=max_c, values_only=True),
        start=1,
    ):
        if _is_row_hidden_extended(target, r_idx):
            continue  # 숨김 행 제외

        arr = [("" if v is None else str(v).strip()) for v in (row or ())]
        while arr and arr[-1] == "":
            arr.pop()
        if any(arr):
            rows.append(arr)

    max_len = max((len(r) for r in rows), default=0)
    rows = [r + [""] * (max_len - len(r)) for r in rows]
    return rows

# ------------------------------------------------------
# 2) pandas 폴백: 전체 행 읽기 (FutureWarning 제거 버전)
# ------------------------------------------------------
def _read_with_pandas_all_rows(file_bytes: bytes) -> List[List[str]]:
    try:
        df = pd.read_excel(BytesIO(file_bytes), engine="openpyxl", header=None, dtype=str)
    except Exception:
        return []
    # applymap → Series.map(경고 제거)
    df = df.apply(lambda s: s.map(lambda x: "" if pd.isna(x) else str(x).strip()))

    # 오른쪽 빈 열 / 아래쪽 빈 행 제거
    while df.shape[1] > 0 and (df.iloc[:, -1] == "").all():
        df = df.iloc[:, :-1]
    while df.shape[0] > 0 and (df.iloc[-1, :] == "").all():
        df = df.iloc[:-1, :]

    return df.values.tolist()

# ------------------------------------------------------
# 3) Shopee 상단 라벨/메타 행 제거
# ------------------------------------------------------
def _strip_shopee_meta_rows(values: list[list[str]]) -> list[list[str]]:
    if not values:
        return values

    def norm_row(row: list[str]) -> list[str]:
        return [str(c or "").strip() for c in row]

    def is_meta_row(row: list[str]) -> bool:
        r = norm_row(row)
        first = (r[0].lower() if r else "")
        if first in {"basic_info", "media_info", "sales_info"}:
            return True
        return any("search_condition" in c.lower() for c in r if c)

    label_pat = re.compile(r"^(et_title_|ps_)", re.IGNORECASE)

    def is_label_row(row: list[str]) -> bool:
        r = [c.lower() for c in norm_row(row) if c]
        if not r:
            return False
        label_like = sum(1 for c in r if label_pat.match(c) or "ps_item_image" in c or "option_" in c or "option." in c)
        ratio = label_like / max(1, len(r))
        return ratio >= 0.6

    v = values[:]
    while v and is_label_row(v[0]):
        v = v[1:]
    if v and is_meta_row(v[0]):
        v = v[1:]
    elif len(v) >= 2 and is_meta_row(v[1]):
        v = [v[0]] + v[2:]
    return v

# ------------------------------------------------------
# 4) 최종 파서
# ------------------------------------------------------
def read_xlsx_values(bio: BytesIO) -> List[List[str]]:
    sanitized = _sanitize_xlsx_remove_sheetviews(bio)
    sanitized.seek(0)
    raw_bytes = sanitized.read()

    try:
        vis = _read_with_openpyxl_visible_only(raw_bytes)
    except Exception:
        vis = []

    need_fallback = (len(vis) == 0) or (len(vis) <= 1 and (len(vis[0]) if vis else 0) <= 1)
    if need_fallback:
        full = _read_with_pandas_all_rows(raw_bytes)
        if len(full) > len(vis) or (full and (not vis or len(full[0]) > len(vis[0]))):
            vis = full

    vis = _strip_shopee_meta_rows(vis)
    return vis

# ------------------------------------------------------
# 5) Google Sheet 쓰기 (RAW + 청크)
# ------------------------------------------------------
def _write_values_to_sheet(sh: gspread.Spreadsheet, tab: str, values: List[List], logs: List[str]) -> None:
    rows = len(values)
    cols = max((len(r) for r in values), default=0)
    logs.append(f"[INFO] {tab}: parsed shape = {rows}x{cols}")

    if rows == 0 or cols == 0:
        logs.append(f"[WARN] {tab}: 입력 데이터가 비어 있어 skip")
        return

    try:
        ws = safe_worksheet(sh, tab)
        with_retry(lambda: ws.clear())
    except Exception:
        ws = with_retry(lambda: sh.add_worksheet(title=tab, rows=max(rows + 10, 100), cols=max(cols + 5, 26)))

    if ws.row_count < rows or ws.col_count < cols:
        with_retry(lambda: ws.resize(rows=rows + 10, cols=cols + 5))

    chunk_rows = int(get_env("UPLOAD_CHUNK_ROWS", "0") or "0")
    if chunk_rows > 0:
        start = 0
        while start < rows:
            end = min(rows, start + chunk_rows)
            start_a1 = rowcol_to_a1(start + 1, 1)
            with_retry(lambda: ws.update(start_a1, values[start:end], raw=True))
            start = end
    else:
        end_a1 = rowcol_to_a1(rows, cols)
        with_retry(lambda: ws.update(f"A1:{end_a1}", values, raw=True))

    logs.append(f"[OK] {tab}: {rows}x{cols} 적용 완료")

# ------------------------------------------------------
# 6) 파일명 → 탭 자동 라우팅
# ------------------------------------------------------
def _target_tab(filename: str) -> Optional[str]:
    low = filename.lower()
    if "basic" in low: return "BASIC"
    if "media" in low: return "MEDIA"
    if "sales" in low: return "SALES"
    return None

# ------------------------------------------------------
# 7) 업로드 반영 엔트리
# ------------------------------------------------------
def apply_uploaded_files(files: dict[str, BytesIO]) -> list[str]:
    """
    Streamlit에서 업로드된 {filename: BytesIO}를 받아
    파일명 규칙(basic/media/sales)에 따라 탭을 초기화하고 RAW+청크로 값을 붙여넣는다.
    """
    logs: List[str] = []
    if not files:
        return ["[WARN] 업로드된 파일이 없습니다."]

    sh = open_sheet_by_env()

    for fname, raw in files.items():
        tab = _target_tab(fname)
        if not tab:
            logs.append(f"[SKIP] 파일명 규칙 불일치: {fname}")
            continue

        try:
            raw.seek(0)
            values = read_xlsx_values(raw)
        except Exception as e:
            logs.append(f"[ERROR] {tab}: {fname} 읽기 실패 → {e}")
            continue

        if len(values) <= 1 and (len(values[0]) if values else 0) <= 1:
            logs.append(f"[WARN] {tab}: 데이터가 비정상적으로 작습니다. (shape={len(values)}x{len(values[0]) if values else 0})")

        try:
            _write_values_to_sheet(sh, tab, values, logs)
        except Exception as e:
            logs.append(f"[ERROR] {tab}: {fname} 반영 실패 → {e}")

    if not any(x.startswith("[OK]") for x in logs):
        logs.append("[WARN] 적용된 탭이 없습니다. 파일명 규칙/시트 권한을 확인하세요.")
    return logs

# ------------------------------------------------------
# 8) Streamlit 업로더용 수집기 (여러 XLSX 동시)
# ------------------------------------------------------
def collect_xlsx_files(files) -> dict[str, BytesIO]:
    """
    Streamlit file_uploader(accept_multiple_files=True)의 반환값을
    {filename: BytesIO} 형태로 변환한다.
    """
    out: dict[str, BytesIO] = {}
    for f in files or []:
        name = getattr(f, "name", "")
        if name.lower().endswith(".xlsx"):
            out[name] = BytesIO(f.read())
    return out
